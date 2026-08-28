# -*- coding: utf-8 -*-
"""Excel (.xlsx) hesabatları — canlı formullarla, üzərinə yazıla bilən."""
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = dict(name="Arial", size=10)
HEAD_FILL = PatternFill("solid", fgColor="E2EFEE")
MATCH_FILL = PatternFill("solid", fgColor="D7F0EC")
LINK_FONT = Font(**ARIAL, color="0563C1", underline="single")
INPUT_FONT = Font(**ARIAL, color="0000FF")     # mavi = əl ilə dəyişilə bilən xana
LEVELS = {2: "dəqiq", 1: "ehtimal", 0: "yox"}


def _safe_name(text: str) -> str:
    return re.sub(r"[^\w\-]", "_", text or "")[:40] or "arma"


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(**ARIAL, bold=True)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _buffer(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def check_report(check, offers):
    """Bir yoxlamanın hesabatı. Maya və hədd əmsalı canlı xanadır —
    Excel-də dəyişəndə marja və status sütunları özü yenilənir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Marja Hesabatı"

    ws["A1"] = f"ARMA Marja Hesabatı — {check['code']}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    ws["A2"] = f"Tarix: {check['created_at']}   Maya: {check['cost']} ₼"
    ws["A2"].font = Font(**ARIAL)
    ws["A3"] = "Hədd əmsalı:"
    ws["A3"].font = Font(**ARIAL)
    ws["B3"] = check["threshold"]
    ws["B3"].font = INPUT_FONT

    _headers(ws, 5, ["Məhsul", "Satıcı", "Kod uyğunluğu", "Qiymət (₼)", "Maya (₼)",
                     "Hədd (₼)", "Marja %", "Status", "Link"])
    r = 6
    for o in offers:
        ws.cell(row=r, column=1, value=o["name"]).font = Font(**ARIAL)
        ws.cell(row=r, column=2, value=o["seller"]).font = Font(**ARIAL)
        ws.cell(row=r, column=3, value=LEVELS.get(o["code_match"], "yox")).font = Font(**ARIAL)
        ws.cell(row=r, column=4, value=o["price"]).font = Font(**ARIAL)
        ws.cell(row=r, column=5, value=check["cost"]).font = INPUT_FONT
        ws.cell(row=r, column=6, value=f"=E{r}*$B$3").font = Font(**ARIAL)
        mp = ws.cell(row=r, column=7, value=f"=IF(E{r}=0,\"\",D{r}/E{r}-1)")
        mp.font = Font(**ARIAL)
        mp.number_format = "0.0%"
        ws.cell(row=r, column=8,
                value=(f'=IF(C{r}="yox","kod uyğun deyil",'
                       f'IF(D{r}<F{r},"aşağı",'
                       f'IF(C{r}="dəqiq","UYĞUN","EHTİMAL — yoxla")))')).font = Font(**ARIAL)
        lk = ws.cell(row=r, column=9, value=o["url"] or "")
        lk.font = LINK_FONT
        if o["url"]:
            lk.hyperlink = o["url"]
        if o["is_match"]:
            for col in range(1, 10):
                ws.cell(row=r, column=col).fill = MATCH_FILL
        for col in (4, 5, 6):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        r += 1

    _widths(ws, [44, 20, 13, 12, 12, 12, 10, 18, 55])
    ws.freeze_panes = "A6"
    return _buffer(wb), f"marja_{_safe_name(check['code'])}_{check['id']}.xlsx"


def links_report(groups, label):
    """Bir gün / bir seçim üçün bütün aktiv linklər."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Linklər"
    ws["A1"] = f"ARMA — uyğun linklər · {label}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)

    _headers(ws, 3, ["Kod", "Maya (₼)", "Qiymət (₼)", "Marja %", "Məhsul", "Link"])
    r = 4
    for grp in groups:
        c = grp["check"]
        for o in grp["offers"]:
            ws.cell(row=r, column=1, value=c["code"]).font = Font(**ARIAL, bold=True)
            ws.cell(row=r, column=2, value=c["cost"]).font = Font(**ARIAL)
            ws.cell(row=r, column=3, value=o["price"]).font = Font(**ARIAL)
            mc = ws.cell(row=r, column=4, value=(o["margin"] or 0) / 100)
            mc.number_format = "0.0%"
            mc.font = Font(**ARIAL)
            ws.cell(row=r, column=5, value=o["name"]).font = Font(**ARIAL)
            lk = ws.cell(row=r, column=6, value=o["url"])
            lk.font = LINK_FONT
            if o["url"]:
                lk.hyperlink = o["url"]
            for col in (2, 3):
                ws.cell(row=r, column=col).number_format = "#,##0.00"
            r += 1

    _widths(ws, [14, 10, 11, 9, 46, 55])
    ws.freeze_panes = "A4"
    return _buffer(wb), f"linkler_{_safe_name(label)}.xlsx"
